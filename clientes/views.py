from django.shortcuts import render, get_object_or_404, redirect
from django.core.paginator import Paginator
from django.db.models import Q
from .forms import ClienteForm, SegmentoForm, AtividadeForm
from .models import Cliente, Segmento, Atividade, ESTADOS
from atividades.models import AtividadeUsuarioCliente
from django.contrib.auth.decorators import login_required
from usuarios.models import Perfil, Chave_Gerenciador
from django.urls import reverse
from django.contrib import messages
from django.http import HttpResponseRedirect
import pandas as pd
from django.http import HttpResponse
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from configuracoes.models import Empresa

@login_required
def cadastrar_clientes(request):
    if request.method == 'POST':
        form = ClienteForm(request.POST)
        if form.is_valid():
            cliente = form.save()  # salva o cliente e guarda o objeto
            # registra a atividade do usuário
            AtividadeUsuarioCliente.objects.create(
                usuario=request.user,
                cliente=cliente,
                descricao="Cadastrou um novo cliente"
            )
            messages.success(request, 'Novo cliente criado com sucesso.')
            return redirect('clientes:listar_clientes')
    else:
        form = ClienteForm()

    return render(request, 'clientes/cadastrar_clientes.html', {'form': form})

@login_required
def listar_clientes(request):
    pesquisa = request.GET.get('pesquisa', '')
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    clientes = Cliente.objects.all().order_by('-id')

    if pesquisa:
        clientes = clientes.filter(
            Q(codigo__icontains=pesquisa) |
            Q(nome_cliente__icontains=pesquisa) |
            Q(nome_fantasia__icontains=pesquisa) |
            Q(email1__icontains=pesquisa) |
            Q(email2__icontains=pesquisa) |
            Q(cnpj_cpf__icontains=pesquisa) |
            Q(telefone1__icontains=pesquisa) |
            Q(telefone2__icontains=pesquisa) |
            Q(celular1__icontains=pesquisa) | 
            Q(celular2__icontains=pesquisa) |
            Q(inscricao_estadual__icontains=pesquisa) |
            Q(inscricao_municipal__icontains=pesquisa)
        )

    if data_inicio:
        clientes = clientes.filter(data_inclusao__gte=data_inicio)
        
    if data_fim:
        clientes = clientes.filter(data_inclusao__lte=data_fim)

    paginator = Paginator(clientes, 5)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    qtd_clientes = Cliente.objects.count()
    qtd_segmentos = Segmento.objects.count()
    qtd_atividades = Atividade.objects.count()
    qtd_estados = Cliente.objects.values('estado_real').distinct().count()
    return render(request, 'clientes/clientes.html', {
        'page_obj': page_obj,
        'pesquisa': pesquisa,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        "qtd_clientes": qtd_clientes,
        "qtd_segmentos": qtd_segmentos,
        "qtd_atividades": qtd_atividades,
        "qtd_estados": qtd_estados
    })

@login_required
def editar_cliente(request, cliente_id):
    cliente = get_object_or_404(Cliente, id=cliente_id)
    if request.method == 'POST':
        form = ClienteForm(request.POST, instance=cliente)
        if form.is_valid():
            cliente = form.save()  # salva o cliente e guarda o objeto
            # registra a atividade do usuário
            AtividadeUsuarioCliente.objects.create(
                usuario=request.user,
                cliente=cliente,
                descricao="Editou um cliente"
            )
            messages.success(request, 'Cliente editado com sucesso.')
            return redirect('clientes:listar_clientes')
    else:
        form = ClienteForm(instance=cliente)
    return render(request, 'clientes/editar_clientes.html', {'form': form, 'cliente': cliente})

@login_required
def excluir_cliente(request, cliente_id):
    cliente = get_object_or_404(Cliente, id=cliente_id)
    cliente.delete()
    messages.success(request, 'Cliente excluído com sucesso.')
    return redirect('clientes:listar_clientes')

@login_required
def ver_cliente(request, cliente_id):
    cliente = get_object_or_404(Cliente, id=cliente_id)
    empresa = Empresa.objects.first()
    return render(request, 'clientes/ver_cliente.html', {'cliente': cliente, "empresa":empresa})

@login_required
def segmentos_atividades(request):
    usuario = request.user
    try:
        perfil = Perfil.objects.get(usuario=usuario)
    except Perfil.DoesNotExist:
        mensagem = 'Perfil não encontrado'
        return render(request, 'principal/erro.html', {'mensagem': mensagem})

    if perfil.tipo not in ['gerenciador', 'supervisor']:
        mensagem = 'Você não tem permissão para acessar esta página.'
        return render(request, 'principal/erro.html', {'mensagem': mensagem})

    segmento = Segmento.objects.all()
    atividade = Atividade.objects.all()
    
    return render(request, 'clientes/segmentos_atividade.html', {
        "atividades": atividade,
        "segmentos": segmento
    })

@login_required
def editar_segmento(request, seg_id):
    segmento = get_object_or_404(Segmento, id=seg_id)
    if request.method == 'POST':
        form = SegmentoForm(request.POST, instance=segmento)
        if form.is_valid():
            form.save()
            return redirect('clientes:segmentos_atividades')  # Ajuste para a URL correta
    else:
        form = SegmentoForm(instance=segmento)
    return render(request, 'clientes/editar_segmento.html', {'form': form, 'segmento': segmento})

@login_required
def excluir_segmento(request, seg_id):
    segmento = get_object_or_404(Segmento, id=seg_id)
    segmento.delete()
    return redirect ('clientes:segmentos_atividades')  

@login_required
def editar_atividade(request, atv_id):
    atividade = get_object_or_404(Atividade, id=atv_id)
    if request.method == 'POST':
        form = AtividadeForm(request.POST, instance=atividade)
        if form.is_valid():
            form.save()
            return HttpResponseRedirect(reverse('clientes:segmentos_atividades') + '?aba=atividades')
    else:
        form = AtividadeForm(instance=atividade)
    return render(request, 'clientes/editar_atividade.html', {'form': form, 'atividade': atividade})

@login_required
def excluir_atividade(request, atv_id):
    atividade = get_object_or_404(Atividade, id=atv_id)
    atividade.delete()
    return HttpResponseRedirect(reverse('clientes:segmentos_atividades') + '?aba=atividades')

@login_required
def cadastrar_segmento(request):
    if request.method == 'POST':
        form = SegmentoForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('clientes:segmentos_atividades')  # Ajuste para a URL correta
    else:
        form = SegmentoForm()
    return render(request, 'clientes/cadastrar_segmento.html', {'form': form})

@login_required
def cadastrar_atividade(request):
    if request.method == 'POST':
        form = AtividadeForm(request.POST)
        if form.is_valid():
            form.save()
            return HttpResponseRedirect(reverse('clientes:segmentos_atividades') + '?aba=atividades')
    else:
        form = AtividadeForm()
    return render(request, 'clientes/cadastrar_atividade.html', {'form': form})

def ajustar_largura_colunas(worksheet, dataframe):
    for i, col in enumerate(dataframe.columns, 1):
        max_length = max(
            dataframe[col].astype(str).map(len).max(),
            len(str(col))
        )
        adjusted_width = max_length + 2
        column_letter = get_column_letter(i)
        worksheet.column_dimensions[column_letter].width = adjusted_width

def download_modelo_clientes(request):
    campos_excluidos = ['id', 'codigo']
    campos_modelo = []
    exemplo = {}

    # ForeignKey: opções do banco
    atividades = list(Atividade.objects.values_list('nome', flat=True))
    segmentos = list(Segmento.objects.values_list('nome', flat=True))

    # Choices fixos
    estados = [sigla for sigla, nome in ESTADOS]

    # Campos que terão dropdowns
    campos_dropdown = {
        'atividade': atividades,
        'segmento': segmentos,
        'estado_real': estados,
        'estado_cobranca': estados,
    }

    # Monta os campos e o exemplo
    for field in Cliente._meta.get_fields():
        if field.concrete and not field.many_to_many and field.name not in campos_excluidos:
            nome_campo = field.name
            campos_modelo.append(nome_campo)

            if nome_campo == 'nome_cliente':
                exemplo[nome_campo] = 'Empresa Exemplo'
            elif nome_campo == 'data_inclusao':
                exemplo[nome_campo] = '00/00/0000'
            elif nome_campo == 'cnpj_cpf':
                exemplo[nome_campo] = '00.000.000/0000-00 ou 000.000.000-00'
            elif nome_campo == 'email1':
                exemplo[nome_campo] = 'exemplo@gmail.com'
            elif nome_campo == 'telefone1' or  nome_campo == 'telefone2' or nome_campo == 'celular1' or  nome_campo == 'celular2':
                exemplo[nome_campo] = '(00) 00000-0000'
            elif nome_campo == 'cep_real' or  nome_campo == 'cep_cobranca':
                exemplo[nome_campo] = '00000-000'

            else:
                exemplo[nome_campo] = ''

    df_modelo = pd.DataFrame([exemplo], columns=campos_modelo)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=modelo_clientes.xlsx'

    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df_modelo.to_excel(writer, sheet_name='Modelo', index=False)
        workbook = writer.book
        sheet = workbook['Modelo']

        ajustar_largura_colunas(sheet, df_modelo)

        # Adiciona validação para campos com dropdown
        for nome_campo, opcoes in campos_dropdown.items():
            if nome_campo in campos_modelo and opcoes:
                col_index = campos_modelo.index(nome_campo) + 1
                col_letra = get_column_letter(col_index)
                val_list = ",".join([str(op).replace(",", " ") for op in opcoes])  # remove vírgula do conteúdo

                dv = DataValidation(type="list", formula1=f'"{val_list}"', allow_blank=True)
                dv.error = 'Escolha uma opção válida.'
                dv.errorTitle = 'Entrada inválida'
                dv.prompt = f'Selecione um {nome_campo}'
                dv.promptTitle = f'{nome_campo.capitalize()}'

                sheet.add_data_validation(dv)
                dv.add(f"{col_letra}2:{col_letra}100")  # aplica nas linhas 2 a 100 (pode ajustar)

    return response

def importar_clientes(request):
    if request.method == 'POST' and request.FILES.get('arquivo'):
        arquivo = request.FILES['arquivo']
        try:
            df = pd.read_excel(arquivo)

            # Função para limpar valores NaN
            def limpar_valor(valor):
                return '' if pd.isna(valor) else valor

            # Campos esperados no arquivo (todos menos id e codigo)
            campos_excluidos = ['id', 'codigo']
            campos_modelo = [f.name for f in Cliente._meta.get_fields()
                             if f.concrete and not f.many_to_many and f.name not in campos_excluidos]

            # Verifica se todas as colunas necessárias estão no arquivo
            faltando = set(campos_modelo) - set(df.columns)
            if faltando:
                messages.error(request, f'Colunas faltando na planilha: {", ".join(faltando)}')
                return redirect('clientes:listar_clientes')

            # Campos obrigatórios baseados no model
            campos_obrigatorios = [
                f.name for f in Cliente._meta.get_fields()
                if getattr(f, 'blank', False) is False and not f.auto_created and f.name not in campos_excluidos
            ]

            # Mapeamento dos estados válidos
            estados_validos = {sigla for sigla, _ in ESTADOS}

            for idx, row in df.iterrows():
                # Validação de campos obrigatórios
                for campo in campos_obrigatorios:
                    if campo not in row or pd.isna(row[campo]):
                        messages.error(request, f"Linha {idx+2}: campo obrigatório '{campo}' faltando. Ignorada.")
                        return redirect('clientes:listar_clientes')

                # Verifica estados
                estado_real = str(row.get('estado_real', '')).strip().upper()
                estado_cobranca = str(row.get('estado_cobranca', '')).strip().upper()
                if estado_real not in estados_validos:
                    messages.error(request, f"Linha {idx+2}: estado_real inválido '{estado_real}'. Ignorada.")
                    return redirect('clientes:listar_clientes')
                if estado_cobranca and estado_cobranca not in estados_validos:
                    messages.error(request, f"Linha {idx+2}: estado_cobranca inválido '{estado_cobranca}'. Ignorada.")
                    return redirect('clientes:listar_clientes')

                # Pega ou cria Atividade
                atividade_nome = str(row.get('atividade', '')).strip()
                atividade_obj = None
                if atividade_nome:
                    atividade_obj, _ = Atividade.objects.get_or_create(nome=atividade_nome)

                # Pega ou cria Segmento
                segmento_nome = str(row.get('segmento', '')).strip()
                segmento_obj = None
                if segmento_nome:
                    segmento_obj, _ = Segmento.objects.get_or_create(nome=segmento_nome)

                # Cria ou atualiza cliente
                cliente, created = Cliente.objects.update_or_create(
                    cnpj_cpf=str(row['cnpj_cpf']).strip(),
                    defaults={
                        'nome_cliente': limpar_valor(row['nome_cliente']),
                        'nome_fantasia': limpar_valor(row.get('nome_fantasia')),
                        'data_inclusao': limpar_valor(row['data_inclusao']),
                        'inscricao_estadual': limpar_valor(row.get('inscricao_estadual')),
                        'inscricao_municipal': limpar_valor(row.get('inscricao_municipal')),
                        'email1': limpar_valor(row['email1']),
                        'email2': limpar_valor(row.get('email2')),
                        'telefone1': limpar_valor(row.get('telefone1')),
                        'telefone2': limpar_valor(row.get('telefone2')),
                        'celular1': limpar_valor(row.get('celular1')),
                        'celular2': limpar_valor(row.get('celular2')),
                        'atividade': atividade_obj,
                        'segmento': segmento_obj,
                        'observacao': limpar_valor(row.get('observacao')),
                        'logradouro_real': limpar_valor(row.get('logradouro_real')),
                        'numero_real': str(limpar_valor(row.get('numero_real')) or ''),
                        'complemento_real': limpar_valor(row.get('complemento_real')),
                        'bairro_real': limpar_valor(row.get('bairro_real')),
                        'cidade_real': limpar_valor(row.get('cidade_real')),
                        'estado_real': estado_real,
                        'cep_real': str(limpar_valor(row.get('cep_real')) or ''),
                        'logradouro_cobranca': limpar_valor(row.get('logradouro_cobranca')),
                        'numero_cobranca': str(limpar_valor(row.get('numero_cobranca')) or ''),
                        'complemento_cobranca': limpar_valor(row.get('complemento_cobranca')),
                        'bairro_cobranca': limpar_valor(row.get('bairro_cobranca')),
                        'cidade_cobranca': limpar_valor(row.get('cidade_cobranca')),
                        'estado_cobranca': estado_cobranca if estado_cobranca else None,
                        'cep_cobranca': str(limpar_valor(row.get('cep_cobranca')) or ''),
                    }
                )

            messages.success(request, 'Planilha importada com sucesso!')

        except Exception as e:
            messages.error(request, f'Erro ao importar planilha: {e}')

        return redirect('clientes:listar_clientes')

    # GET
    return render(request, 'clientes/listar_clientes.html')

